"""Parse BTG Pactual credit card fatura XLSX exports into transactions."""

from __future__ import annotations

import re
from datetime import date, datetime, timedelta
from decimal import Decimal
from pathlib import Path

from models import Transaction

try:
    import openpyxl
except ImportError:  # pragma: no cover - exercised when openpyxl is missing
    openpyxl = None

_COL_LABEL = 2
_COL_VALUE = 4
_COL_DATE = 2
_COL_DESCRIPTION = 3
_COL_AMOUNT = 5
_COL_TYPE = 6

_PERIOD_RE = re.compile(
    r"(\d{2})/(\d{2})\s+até\s+(\d{2})/(\d{2})",
    re.IGNORECASE,
)
_DUE_DATE_RE = re.compile(r"(\d{2})/(\d{2})")
_FILENAME_DATE_RE = re.compile(r"(\d{4})-(\d{2})-(\d{2})")
_EXCEL_EPOCH = date(1899, 12, 30)


def parse_fatura(path: Path | str) -> tuple[str, str, list[Transaction]]:
    """Return (period_start, period_end, transactions) from a BTG fatura XLSX export."""
    if openpyxl is None:
        raise ImportError(
            "Reading .xlsx files requires openpyxl. Install it with: pip install openpyxl"
        )

    workbook = openpyxl.load_workbook(str(path), data_only=True)
    sheet = workbook.active
    due_year, due_month = _extract_due_date(sheet, Path(path))
    period_start, period_end = _extract_period(sheet, due_year, due_month)
    transactions = _extract_transactions(sheet)
    return period_start, period_end, transactions


def _cell_str(sheet: openpyxl.worksheet.worksheet.Worksheet, row: int, col: int) -> str:
    value = sheet.cell(row, col).value
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _extract_due_date(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    path: Path,
) -> tuple[int, int]:
    for row in range(1, sheet.max_row + 1):
        if _cell_str(sheet, row, _COL_LABEL) != "Vencimento":
            continue
        match = _DUE_DATE_RE.search(_cell_str(sheet, row, _COL_VALUE))
        if match:
            day, month = match.groups()
            year = _infer_year_from_filename(path) or datetime.now().year
            return year, int(month)

    filename_match = _FILENAME_DATE_RE.search(path.name)
    if filename_match:
        year, month, _ = filename_match.groups()
        return int(year), int(month)

    now = datetime.now()
    return now.year, now.month


def _infer_year_from_filename(path: Path) -> int | None:
    match = _FILENAME_DATE_RE.search(path.name)
    if not match:
        return None
    return int(match.group(1))


def _extract_period(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    due_year: int,
    due_month: int,
) -> tuple[str, str]:
    for row in range(1, sheet.max_row + 1):
        if _cell_str(sheet, row, _COL_LABEL) != "Período de Compras":
            continue
        period_text = _cell_str(sheet, row, _COL_VALUE)
        match = _PERIOD_RE.search(period_text)
        if not match:
            break
        start_day, start_month, end_day, end_month = match.groups()
        start_year = due_year if int(start_month) <= due_month else due_year - 1
        end_year = due_year if int(end_month) <= due_month else due_year - 1
        return (
            f"{start_year}{start_month}{start_day}000000",
            f"{end_year}{end_month}{end_day}235959",
        )
    return "?", "?"


def _extract_transactions(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
) -> list[Transaction]:
    transactions: list[Transaction] = []
    in_table = False

    for row in range(1, sheet.max_row + 1):
        date_label = _cell_str(sheet, row, _COL_DATE)
        if date_label == "Data":
            in_table = True
            continue
        if not in_table:
            continue

        parsed_date = _parse_date(sheet.cell(row, _COL_DATE).value)
        if parsed_date is None:
            continue

        description = _cell_str(sheet, row, _COL_DESCRIPTION)
        if not description:
            continue

        amount = _cell_decimal(sheet, row, _COL_AMOUNT)
        if amount == 0:
            continue

        trn_type = _cell_str(sheet, row, _COL_TYPE) or "UNKNOWN"

        transactions.append(
            Transaction(
                trn_type=trn_type,
                date=parsed_date,
                amount=-abs(amount),
                memo=trn_type,
                name=description,
            )
        )

    return transactions


def _parse_date(value: object) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y%m%d%H%M%S")
    if isinstance(value, date):
        return value.strftime("%Y%m%d") + "000000"
    if isinstance(value, (int, float)):
        parsed = _EXCEL_EPOCH + timedelta(days=int(value))
        return parsed.strftime("%Y%m%d") + "000000"

    text = str(value).strip()
    for fmt in ("%d/%m/%Y", "%d/%m/%Y %H:%M", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).strftime("%Y%m%d%H%M%S")
        except ValueError:
            continue
    return None


def _cell_decimal(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    row: int,
    col: int,
) -> Decimal:
    value = sheet.cell(row, col).value
    if value is None or value == "":
        return Decimal("0")
    if isinstance(value, (int, float)):
        return Decimal(str(value)).quantize(Decimal("0.01"))
    normalized = str(value).strip().replace(".", "").replace(",", ".")
    return Decimal(normalized).quantize(Decimal("0.01"))
